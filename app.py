import streamlit as st
from supabase import create_client
from dotenv import load_dotenv
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import os
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import openpyxl
import re
import yfinance as yf

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

@st.cache_resource
def init_supabase():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

def get_supabase_auth():
    """Devuelve cliente Supabase con token del usuario autenticado."""
    client = init_supabase()
    if "access_token" in st.session_state and st.session_state.access_token:
        client.postgrest.auth(st.session_state.access_token)
    return client

def login_page():
    if "user" not in st.session_state:
        st.session_state.user = None
    if "access_token" not in st.session_state:
        st.session_state.access_token = None

    if st.session_state.user:
        return True

    st.title("💰 Money Magnet")
    st.subheader("Iniciar sesión")

    email = st.text_input("Email")
    password = st.text_input("Contraseña", type="password")

    if st.button("Entrar", type="primary"):
        try:
            supabase = init_supabase()
            response = supabase.auth.sign_in_with_password(
                credentials={"email": email, "password": password}
            )
            st.session_state.user = response.user
            st.session_state.access_token = response.session.access_token
            st.rerun()
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

    return False

def get_user_id():
    return st.session_state.user.id

def procesar_xlsx(archivo, nombre_cuenta="Euros"):
    df = pd.read_excel(archivo)
    columnas_esperadas = [
        "Según un período", "Cuentas", "Categoría",
        "Subcategorías", "Nota", "EUR", "Ingreso/Gasto", "Descripción"
    ]
    columnas_faltantes = [c for c in columnas_esperadas if c not in df.columns]
    if columnas_faltantes:
        raise ValueError(f"Columnas faltantes en el archivo: {columnas_faltantes}")
    df = df[df["Cuentas"] == nombre_cuenta].copy()
    df = df.rename(columns={
        "Según un período": "fecha_gasto",
        "Cuentas": "cuenta",
        "Categoría": "categoria_consumo",
        "Subcategorías": "sub_categoria",
        "Nota": "consumo",
        "EUR": "monto",
        "Ingreso/Gasto": "tipo",
        "Descripción": "descripcion"
    })
    df = df[["fecha_gasto", "cuenta", "categoria_consumo",
             "sub_categoria", "consumo", "monto", "tipo", "descripcion"]]
    df["tipo"] = df["tipo"].replace("Gastos", "Gasto")
    df["fecha_gasto"] = pd.to_datetime(df["fecha_gasto"]).dt.date
    df["sub_categoria"] = df["sub_categoria"].fillna("")
    df["consumo"] = df["consumo"].fillna("")
    df["descripcion"] = df["descripcion"].fillna("")
    return df

def sincronizar(df, supabase, user_id):
    supabase.table("gastos").delete().eq("user_id", user_id).execute()
    registros = df.to_dict(orient="records")
    registros_str = []
    for r in registros:
        r["fecha_gasto"] = str(r["fecha_gasto"])
        r["monto"] = float(r["monto"])
        r["user_id"] = user_id
        registros_str.append(r)
    for i in range(0, len(registros_str), 500):
        lote = registros_str[i:i + 500]
        supabase.table("gastos").insert(lote).execute()
    return len(registros_str)

@st.cache_data(ttl=300)
def get_todos_gastos(_supabase, user_id):
    todos = []
    page_size = 1000
    offset = 0
    while True:
        result = _supabase.table("gastos")\
            .select("fecha_gasto, categoria_consumo, consumo, monto, tipo")\
            .eq("user_id", user_id)\
            .range(offset, offset + page_size - 1)\
            .execute()
        if not result.data:
            break
        todos.extend(result.data)
        offset += page_size
    if not todos:
        return pd.DataFrame()
    df = pd.DataFrame(todos)
    df["fecha_gasto"] = pd.to_datetime(df["fecha_gasto"])
    df["importe"] = df.apply(
        lambda r: r["monto"] if r["tipo"] == "Ingreso" else -r["monto"], axis=1
    )
    df["anio"] = df["fecha_gasto"].dt.year
    df["mes"] = df["fecha_gasto"].dt.month
    df["mes_anio"] = df["fecha_gasto"].dt.to_period("M")
    return df

def calcular_stats_gastos_6m(_supabase, user_id):
    """Balance neto mensual de los últimos 6 meses completos: media, mediana, std, máx/mín."""
    df = get_todos_gastos(_supabase, user_id)
    if df.empty:
        return None

    mes_actual = pd.Timestamp.today().to_period('M')
    balances = df.groupby('mes_anio')['importe'].sum()
    balances = balances[balances.index < mes_actual].sort_index().tail(6)

    if balances.empty:
        return None

    meses_es = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

    def nombre_mes(periodo):
        return f"{meses_es[periodo.month]} {periodo.year}"

    return {
        'media': balances.mean(),
        'mediana': balances.median(),
        'std': balances.std() if len(balances) > 1 else 0.0,
        'max_val': balances.max(),
        'max_mes': nombre_mes(balances.idxmax()),
        'min_val': balances.min(),
        'min_mes': nombre_mes(balances.idxmin()),
        'n_meses': len(balances)
    }

def get_gastos_mes(supabase, year, month, user_id):
    inicio = date(year, month, 1)
    fin = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    result = supabase.table("gastos")\
        .select("categoria_consumo, monto, tipo")\
        .eq("user_id", user_id)\
        .gte("fecha_gasto", str(inicio))\
        .lt("fecha_gasto", str(fin))\
        .execute()
    return pd.DataFrame(result.data) if result.data else pd.DataFrame()

def get_presupuestos_mes(supabase, year, month, user_id):
    inicio = date(year, month, 1)
    result = supabase.table("presupuestos")\
        .select("categoria_consumo, monto")\
        .eq("user_id", user_id)\
        .eq("fecha", str(inicio))\
        .execute()
    return pd.DataFrame(result.data) if result.data else pd.DataFrame()

@st.cache_data(ttl=300)
def get_categorias_usuario(_supabase, user_id):
    """Devuelve categorías únicas de presupuestos del usuario."""
    result = _supabase.table("presupuestos")\
        .select("categoria_consumo")\
        .eq("user_id", user_id)\
        .execute()
    return sorted(set(r["categoria_consumo"] for r in (result.data or [])))

def get_presupuestos_anio(supabase, year, user_id):
    """Suma de presupuestos por categoría para el año completo (todas las filas cargadas)."""
    inicio = date(year, 1, 1)
    fin = date(year + 1, 1, 1)
    result = supabase.table("presupuestos")\
        .select("categoria_consumo, monto")\
        .eq("user_id", user_id)\
        .gte("fecha", str(inicio))\
        .lt("fecha", str(fin))\
        .execute()
    if not result.data:
        return pd.DataFrame(columns=["categoria_consumo", "monto"])
    df = pd.DataFrame(result.data)
    return df.groupby("categoria_consumo")["monto"].sum().reset_index()


@st.cache_data(ttl=300)
def get_balance_app(_supabase, user_id):
    todos = []
    offset = 0
    while True:
        result = _supabase.table("gastos")\
            .select("monto, tipo")\
            .eq("user_id", user_id)\
            .range(offset, offset + 999)\
            .execute()
        if not result.data:
            break
        todos.extend(result.data)
        offset += 1000
    if not todos:
        return 0
    df = pd.DataFrame(todos)
    df["importe"] = df.apply(
        lambda r: r["monto"] if r["tipo"] == "Ingreso" else -r["monto"], axis=1
    )
    return df["importe"].sum()

@st.cache_data(ttl=300)
def get_saldos_actuales(_supabase, user_id):
    result = _supabase.table("saldos_bancarios")\
        .select("banco, monto, fecha_registro")\
        .eq("user_id", user_id)\
        .order("fecha_registro", desc=True)\
        .execute()
    if not result.data:
        return pd.DataFrame(), None
    df = pd.DataFrame(result.data)
    ultima_fecha = df["fecha_registro"].max()
    df_actual = df[df["fecha_registro"] == ultima_fecha][["banco", "monto"]].copy()
    return df_actual, ultima_fecha

def guardar_saldos(supabase, saldos_dict, user_id):
    hoy = str(date.today())
    registros = [
        {"banco": banco, "monto": float(monto), "fecha_registro": hoy, "user_id": user_id}
        for banco, monto in saldos_dict.items()
        if banco.strip()
    ]
    if registros:
        supabase.table("saldos_bancarios").insert(registros).execute()

def widget_saldos_inline(supabase, user_id):
    df_saldos, ultima_fecha = get_saldos_actuales(supabase, user_id)

    st.divider()
    st.subheader("💳 ¿Actualizás tus saldos bancarios?")
    if ultima_fecha:
        st.caption(f"Últimos saldos registrados: {ultima_fecha}")

    if "saldos_temp" not in st.session_state:
        if not df_saldos.empty:
            st.session_state.saldos_temp = dict(
                zip(df_saldos["banco"], df_saldos["monto"])
            )
        else:
            st.session_state.saldos_temp = {}

    bancos = list(st.session_state.saldos_temp.keys())
    for banco in bancos:
        col1, col2 = st.columns([2, 1])
        with col1:
            st.text(banco)
        with col2:
            nuevo_monto = st.number_input(
                f"€ {banco}",
                value=float(st.session_state.saldos_temp[banco]),
                step=0.01,
                label_visibility="collapsed",
                key=f"inline_{banco}"
            )
            st.session_state.saldos_temp[banco] = nuevo_monto

    col_no, col_si = st.columns([1, 1])
    with col_no:
        if st.button("Ahora no", key="sync_no"):
            st.session_state.mostrar_saldos_post_sync = False
            st.rerun()
    with col_si:
        if st.button("💾 Guardar saldos", type="primary", key="sync_si"):
            guardar_saldos(supabase, st.session_state.saldos_temp, user_id)
            get_saldos_actuales.clear()
            st.session_state.mostrar_saldos_post_sync = False
            st.session_state.pop("saldos_temp", None)
            st.success("✅ Saldos guardados correctamente")
            st.rerun()

CATEGORIAS_OTROS = [
    "Education", "Other", "Impuesto Bancario", "Clothing",
    "Gifts", "Technology", "Payment", "Medical", "Tramites",
    "Visa", "Hacienda"
]

def pagina_dashboard(supabase, user_id):
    st.title("💰 Money Magnet")

    hoy = date.today()
    if "mes_offset" not in st.session_state:
        st.session_state.mes_offset = 0

    mes_actual = hoy + relativedelta(months=st.session_state.mes_offset)
    year, month = mes_actual.year, mes_actual.month
    nombre_mes = mes_actual.strftime("%B %Y").capitalize()

    col_izq, col_centro, col_der, col_hoy = st.columns([1, 2, 1, 1])
    with col_izq:
        if st.button("◀ Mes anterior"):
            st.session_state.mes_offset -= 1
            st.rerun()
    with col_centro:
        st.markdown(f"<h3 style='text-align:center'>📅 {nombre_mes}</h3>",
                    unsafe_allow_html=True)
    with col_der:
        if st.button("Mes siguiente ▶"):
            st.session_state.mes_offset += 1
            st.rerun()
    with col_hoy:
        if st.session_state.mes_offset != 0:
            if st.button("🏠 Hoy"):
                st.session_state.mes_offset = 0
                st.rerun()

    with st.spinner("Cargando datos..."):
        df_gastos = get_gastos_mes(supabase, year, month, user_id)
        df_presupuestos = get_presupuestos_mes(supabase, year, month, user_id)

    if df_gastos.empty and df_presupuestos.empty:
        st.warning("No hay datos para este mes.")
        return

    if not df_gastos.empty:
        df_gastos["importe"] = df_gastos.apply(
            lambda r: r["monto"] if r["tipo"] == "Ingreso" else -r["monto"], axis=1
        )
        total_ingresos = df_gastos[df_gastos["tipo"] == "Ingreso"]["monto"].sum()
        total_gastado = df_gastos[df_gastos["tipo"] == "Gasto"]["monto"].sum()
        balance = total_ingresos - total_gastado
    else:
        total_ingresos = total_gastado = balance = 0

    presupuesto_total = df_presupuestos["monto"].sum() if not df_presupuestos.empty else 0

    st.divider()
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("💸 Total Gastado", f"€{total_gastado:,.2f}")
    k2.metric("💰 Total Ingresos", f"€{total_ingresos:,.2f}")
    k3.metric("⚖️ Balance", f"€{balance:,.2f}")
    k4.metric("🎯 Presupuesto Neto", f"€{presupuesto_total:,.2f}")

    st.divider()
    st.subheader("📊 Detalle por Categoría")
    ocultar_cero = st.toggle("Ocultar categorías sin presupuesto (€0)", value=False)

    if not df_gastos.empty:
        real_cat = df_gastos.groupby("categoria_consumo")["importe"].sum().reset_index()
        real_cat.columns = ["categoria_consumo", "real"]
    else:
        real_cat = pd.DataFrame(columns=["categoria_consumo", "real"])

    if not df_presupuestos.empty:
        df_tabla = pd.merge(df_presupuestos, real_cat, on="categoria_consumo", how="left")
        df_tabla["real"] = df_tabla["real"].fillna(0)
    else:
        df_tabla = real_cat.copy()
        df_tabla["monto"] = 0

    df_tabla.columns = ["Categoría", "Presupuesto", "Real"]
    df_tabla["Diferencia"] = df_tabla["Real"] - df_tabla["Presupuesto"]

    def semaforo(row):
        if row["Real"] == 0 and row["Presupuesto"] != 0:
            return "🟡"
        elif row["Presupuesto"] < 0 and row["Real"] < row["Presupuesto"]:
            return "🔴"
        elif row["Presupuesto"] > 0 and row["Real"] < row["Presupuesto"]:
            return "🔴"
        else:
            return "🟢"

    df_tabla["Estado"] = df_tabla.apply(semaforo, axis=1)

    mask_otros = df_tabla["Categoría"].isin(CATEGORIAS_OTROS)
    df_principales = df_tabla[~mask_otros].copy()
    df_otros = df_tabla[mask_otros].copy()

    if ocultar_cero:
        df_principales = df_principales[df_principales["Presupuesto"] != 0]

    df_principales = df_principales.reindex(
        df_principales["Real"].abs().sort_values(ascending=False).index
    )

    if not df_otros.empty:
        fila_otros = pd.DataFrame([{
            "Categoría": "Otras Categorías ℹ️",
            "Presupuesto": df_otros["Presupuesto"].sum(),
            "Real": df_otros["Real"].sum(),
            "Diferencia": df_otros["Diferencia"].sum(),
            "Estado": "—"
        }])
        df_final = pd.concat([df_principales, fila_otros], ignore_index=True)
    else:
        df_final = df_principales

    df_mostrar = df_final.copy()
    df_mostrar["Presupuesto"] = df_mostrar["Presupuesto"].apply(lambda x: f"€{x:,.2f}")
    df_mostrar["Real"] = df_mostrar["Real"].apply(lambda x: f"€{x:,.2f}")
    df_mostrar["Diferencia"] = df_mostrar["Diferencia"].apply(lambda x: f"€{x:,.2f}")
    st.dataframe(df_mostrar, use_container_width=True, hide_index=True)

    if not df_otros.empty:
        cats_lista = ", ".join(sorted(df_otros["Categoría"].tolist()))
        st.caption(f"ℹ️ Otras Categorías incluye: {cats_lista}")

def pagina_bancos(supabase, user_id):
    st.title("💳 Saldos Bancarios")

    balance_app = get_balance_app(supabase, user_id)
    df_saldos, ultima_fecha = get_saldos_actuales(supabase, user_id)

    if "saldos_edit" not in st.session_state:
        if not df_saldos.empty:
            st.session_state.saldos_edit = dict(
                zip(df_saldos["banco"], df_saldos["monto"])
            )
        else:
            st.session_state.saldos_edit = {}

    if ultima_fecha:
        st.caption(f"Últimos saldos guardados: {ultima_fecha}")

    st.divider()
    st.markdown(f"📱 **Balance Money Magnet:** €{balance_app:,.2f}")
    st.divider()

    bancos = list(st.session_state.saldos_edit.keys())
    for banco in bancos:
        col1, col2, col3 = st.columns([3, 2, 1])
        with col1:
            st.text(banco)
        with col2:
            nuevo_monto = st.number_input(
                f"€",
                value=float(st.session_state.saldos_edit[banco]),
                step=0.01,
                label_visibility="collapsed",
                key=f"edit_{banco}"
            )
            st.session_state.saldos_edit[banco] = nuevo_monto
        with col3:
            if st.button("🗑️", key=f"del_{banco}", help=f"Eliminar {banco}"):
                del st.session_state.saldos_edit[banco]
                st.rerun()

    st.divider()

    with st.expander("➕ Agregar banco"):
        nuevo_banco = st.text_input("Nombre del banco", key="nuevo_banco_nombre")
        nuevo_monto_banco = st.number_input("Monto (€)", value=0.0, step=0.01,
                                             key="nuevo_banco_monto")
        if st.button("Añadir"):
            if nuevo_banco.strip():
                st.session_state.saldos_edit[nuevo_banco.strip()] = nuevo_monto_banco
                st.rerun()

    total_bancos = sum(st.session_state.saldos_edit.values())
    diferencia = balance_app - total_bancos

    col_t1, col_t2, col_t3 = st.columns(3)
    col_t1.metric("💰 Total bancos", f"€{total_bancos:,.2f}")
    col_t2.metric("📱 Balance app", f"€{balance_app:,.2f}")
    if abs(diferencia) <= 0.01:
        col_t3.metric("⚖️ Diferencia", "€0.00 ✅")
    else:
        col_t3.metric("⚖️ Diferencia", f"€{diferencia:,.2f} ⚠️")

    st.divider()
    if st.button("💾 Guardar saldos", type="primary"):
        guardar_saldos(supabase, st.session_state.saldos_edit, user_id)
        get_saldos_actuales.clear()
        st.session_state.pop("saldos_edit", None)
        st.success("✅ Saldos guardados correctamente")
        st.rerun()

    st.divider()
    with st.expander("📋 Ver historial de saldos"):
        result = supabase.table("saldos_bancarios")\
            .select("banco, monto, fecha_registro")\
            .eq("user_id", user_id)\
            .order("fecha_registro", desc=True)\
            .execute()
        if result.data:
            df_hist = pd.DataFrame(result.data)
            fechas = sorted(df_hist["fecha_registro"].unique(), reverse=True)
            fecha_sel = st.selectbox("Fecha", fechas)
            df_fecha = df_hist[df_hist["fecha_registro"] == fecha_sel][["banco", "monto"]]
            df_fecha.columns = ["Banco", "Monto (€)"]
            df_fecha["Monto (€)"] = df_fecha["Monto (€)"].apply(lambda x: f"€{x:,.2f}")
            st.dataframe(df_fecha, use_container_width=True, hide_index=True)
        else:
            st.info("Sin historial disponible")

def pagina_historico(supabase, user_id):
    st.title("📈 Histórico")

    with st.spinner("Cargando datos históricos..."):
        df = get_todos_gastos(supabase, user_id)

    if df.empty:
        st.warning("No hay datos disponibles.")
        return

    anios_disponibles = sorted(df["anio"].unique(), reverse=True)

    st.subheader("📊 Ingresos vs Gastos por Mes")
    anio_sel = st.selectbox("Año", anios_disponibles, index=0, key="anio_barras")

    df_anio = df[df["anio"] == anio_sel].copy()
    meses_nombres = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                     7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

    df_barras = df_anio.groupby(["mes", "tipo"])["monto"].sum().reset_index()
    df_barras["mes_nombre"] = df_barras["mes"].map(meses_nombres)
    df_barras = df_barras.sort_values("mes")

    fig_barras = px.bar(
        df_barras, x="mes_nombre", y="monto", color="tipo", barmode="group",
        color_discrete_map={"Ingreso": "#82c9a0", "Gasto": "#e8968a"},
        labels={"monto": "€", "mes_nombre": "Mes", "tipo": ""},
        title=f"Ingresos vs Gastos — {anio_sel}"
    )
    fig_barras.update_layout(legend_title_text="")
    st.plotly_chart(fig_barras, use_container_width=True)

    st.divider()
    st.subheader("📈 Balance")

    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        vista_balance = st.radio(
            "Vista",
            ["Cascada mensual", "Balance mensual", "Balance acumulado"],
            horizontal=True
        )
    anios_rango = sorted(df["anio"].unique())
    with col2:
        anio_desde = st.selectbox("Desde año", anios_rango, index=0, key="desde_anio")
    with col3:
        anio_hasta = st.selectbox("Hasta año", anios_rango,
                                   index=len(anios_rango)-1, key="hasta_anio")

    df_rango = df[(df["anio"] >= anio_desde) & (df["anio"] <= anio_hasta)].copy()
    df_bal = df_rango.groupby("mes_anio")["importe"].sum().reset_index()
    df_bal = df_bal.sort_values("mes_anio")
    df_bal["etiqueta"] = df_bal["mes_anio"].astype(str)
    df_bal["acumulado"] = df_bal["importe"].cumsum()

    if vista_balance == "Cascada mensual":
        fig = go.Figure(go.Waterfall(
            orientation="v", measure=["relative"] * len(df_bal),
            x=df_bal["etiqueta"], y=df_bal["importe"],
            connector={"line": {"color": "rgba(150,150,150,0.3)"}},
            increasing={"marker": {"color": "#82c9a0"}},
            decreasing={"marker": {"color": "#e8968a"}},
            hovertemplate="%{x}<br>Δ mes: €%{y:,.2f}<extra></extra>"
        ))
        fig.update_layout(title="Cascada de balance mensual",
                          xaxis_title="Mes", yaxis_title="€", showlegend=False)
    elif vista_balance == "Balance mensual":
        fig = go.Figure(go.Bar(
            x=df_bal["etiqueta"], y=df_bal["importe"],
            marker_color=["#82c9a0" if v >= 0 else "#e8968a" for v in df_bal["importe"]],
            hovertemplate="%{x}<br>Balance: €%{y:,.2f}<extra></extra>"
        ))
        fig.update_layout(title="Balance neto por mes",
                          xaxis_title="Mes", yaxis_title="€", showlegend=False)
        fig.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
    else:
        fig = go.Figure(go.Scatter(
            x=df_bal["etiqueta"], y=df_bal["acumulado"],
            mode="lines+markers",
            line=dict(color="#3498db", width=2.5), marker=dict(size=6),
            hovertemplate="%{x}<br>Acumulado: €%{y:,.2f}<extra></extra>"
        ))
        fig.update_layout(title="Balance acumulado",
                          xaxis_title="Mes", yaxis_title="€", showlegend=False)
        fig.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)

    st.plotly_chart(fig, use_container_width=True)

def pagina_detalle(supabase, user_id):
    st.title("🔍 Detalle de Transacciones")

    with st.spinner("Cargando datos..."):
        df = get_todos_gastos(supabase, user_id)

    if df.empty:
        st.warning("No hay datos disponibles.")
        return

    anios_disponibles = sorted(df["anio"].unique(), reverse=True)
    anio_default_idx = (
        anios_disponibles.index(date.today().year)
        if date.today().year in anios_disponibles else 0
    )

    meses_nombres = {0:"Todos",1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",
                     5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",
                     9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

    col1, col2 = st.columns(2)
    with col1:
        anio_sel = st.selectbox("Año", anios_disponibles, index=anio_default_idx)
    with col2:
        mes_sel = st.selectbox("Mes", list(meses_nombres.values()))

    df_filtrado = df[df["anio"] == anio_sel].copy()
    if mes_sel != "Todos":
        mes_num = [k for k, v in meses_nombres.items() if v == mes_sel][0]
        df_filtrado = df_filtrado[df_filtrado["mes"] == mes_num]

    col_r1, col_r2 = st.columns(2)
    col_r1.metric("📋 Registros", f"{len(df_filtrado):,}")
    col_r2.metric("💰 Balance filtrado", f"€{df_filtrado['importe'].sum():,.2f}")

    st.divider()

    meses_col = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                 7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
    meses_presentes = sorted(df_filtrado["mes"].unique())

    pivot = df_filtrado.groupby(["categoria_consumo", "mes"])["importe"].sum().unstack(level="mes")
    pivot.columns = [meses_col[m] for m in pivot.columns]
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total")

    def fmt(x):
        if pd.isna(x) or x == 0:
            return ""
        return f"€{x:,.2f}"

    pivot_fmt = pivot.map(fmt)
    pivot_fmt.index.name = "Categoría"
    pivot_fmt = pivot_fmt.reset_index()

    altura = (len(pivot_fmt) + 1) * 35 + 10
    st.dataframe(pivot_fmt, use_container_width=True, hide_index=True, height=altura)

def pagina_proyeccion(supabase, user_id):
    anio = date.today().year
    fecha_inicio = f"{anio}-01-01"
    fecha_fin = f"{anio}-12-31"

    st.title(f"🔮 Proyección Anual {anio}")

    with st.spinner("Calculando proyección..."):
        todos_hist = []
        offset = 0
        while True:
            result = supabase.table("gastos")\
                .select("monto, tipo, fecha_gasto")\
                .eq("user_id", user_id)\
                .lt("fecha_gasto", fecha_inicio)\
                .range(offset, offset + 999)\
                .execute()
            if not result.data:
                break
            todos_hist.extend(result.data)
            offset += 1000

        saldo_inicial = 0
        if todos_hist:
            df_hist = pd.DataFrame(todos_hist)
            saldo_inicial = df_hist.apply(
                lambda r: r["monto"] if r["tipo"] == "Ingreso" else -r["monto"], axis=1
            ).sum()

        result_real = supabase.table("gastos")\
            .select("fecha_gasto, monto, tipo")\
            .eq("user_id", user_id)\
            .gte("fecha_gasto", fecha_inicio)\
            .lte("fecha_gasto", fecha_fin)\
            .execute()

        result_presup = supabase.table("presupuestos")\
            .select("fecha, monto")\
            .eq("user_id", user_id)\
            .gte("fecha", fecha_inicio)\
            .lte("fecha", fecha_fin)\
            .execute()

    meses = list(range(1, 13))
    meses_nombres = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                     7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

    balance_real_mes = {m: None for m in meses}
    if result_real.data:
        df_real = pd.DataFrame(result_real.data)
        df_real["fecha_gasto"] = pd.to_datetime(df_real["fecha_gasto"])
        df_real["mes"] = df_real["fecha_gasto"].dt.month
        df_real["importe"] = df_real.apply(
            lambda r: r["monto"] if r["tipo"] == "Ingreso" else -r["monto"], axis=1
        )
        for mes, grupo in df_real.groupby("mes"):
            balance_real_mes[mes] = grupo["importe"].sum()

    balance_presup_mes = {m: 0 for m in meses}
    if result_presup.data:
        df_presup = pd.DataFrame(result_presup.data)
        df_presup["fecha"] = pd.to_datetime(df_presup["fecha"])
        df_presup["mes"] = df_presup["fecha"].dt.month
        for mes, grupo in df_presup.groupby("mes"):
            balance_presup_mes[mes] = grupo["monto"].sum()

    filas = []
    saldo_real = saldo_inicial
    saldo_teorico = saldo_inicial
    mes_actual = date.today().month

    for mes in meses:
        es_real = balance_real_mes[mes] is not None
        balance_r = balance_real_mes[mes] if es_real else None
        balance_p = balance_presup_mes[mes]

        if es_real:
            saldo_real += balance_r
            saldo_real_final = saldo_real
        else:
            saldo_real_final = None

        saldo_teorico += balance_p
        saldo_teorico_final = saldo_teorico

        filas.append({
            "Mes": meses_nombres[mes],
            "Balance Real": balance_r,
            "Saldo Real": saldo_real_final,
            "Presupuesto": balance_p,
            "Saldo Teórico": saldo_teorico_final,
            "es_real": es_real,
            "mes_num": mes
        })

    df_tabla = pd.DataFrame(filas)

    saldo_actual = df_tabla[df_tabla["Saldo Real"].notna()]["Saldo Real"].iloc[-1]
    saldo_dic = df_tabla[df_tabla["mes_num"] == 12]["Saldo Teórico"].iloc[0]
    diferencia = saldo_dic - saldo_actual

    k1, k2, k3 = st.columns(3)
    k1.metric("💰 Saldo actual", f"€{saldo_actual:,.2f}")
    k2.metric("🔮 Proyección diciembre", f"€{saldo_dic:,.2f}")
    k3.metric("📈 Diferencia", f"€{diferencia:,.2f}")

    st.divider()

    fig = go.Figure()

    df_real_plot = df_tabla[df_tabla["Saldo Real"].notna()]
    fig.add_trace(go.Scatter(
        x=df_real_plot["Mes"], y=df_real_plot["Saldo Real"],
        mode="lines+markers", name="Saldo Real",
        line=dict(color="#3498db", width=2.5), marker=dict(size=8),
        hovertemplate="%{x}<br>Saldo Real: €%{y:,.2f}<extra></extra>"
    ))

    fig.add_trace(go.Scatter(
        x=df_tabla["Mes"], y=df_tabla["Saldo Teórico"],
        mode="lines+markers", name="Saldo Teórico",
        line=dict(color="#95a5a6", width=2, dash="dot"), marker=dict(size=6),
        hovertemplate="%{x}<br>Saldo Teórico: €%{y:,.2f}<extra></extra>"
    ))

    mes_corte = meses_nombres[mes_actual]
    fig.add_shape(type="line", x0=mes_corte, x1=mes_corte, y0=0, y1=1,
                  xref="x", yref="paper", line=dict(color="orange", width=2, dash="dash"))
    fig.add_annotation(x=mes_corte, y=1, xref="x", yref="paper", text="▲ Hoy",
                       showarrow=False, font=dict(color="orange", size=12), yanchor="bottom")
    fig.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.3)
    fig.update_layout(
        title=f"Proyección de Saldo {anio}", xaxis_title="Mes", yaxis_title="€",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified"
    )
    st.plotly_chart(fig, use_container_width=True)

    st.divider()
    st.subheader("📋 Detalle mes a mes")

    df_mostrar = df_tabla.copy()
    df_mostrar["Balance Real"] = df_mostrar["Balance Real"].apply(
        lambda x: f"€{x:,.2f}" if x is not None else "—")
    df_mostrar["Saldo Real"] = df_mostrar["Saldo Real"].apply(
        lambda x: f"€{x:,.2f}" if x is not None else "—")
    df_mostrar["Presupuesto"] = df_mostrar["Presupuesto"].apply(lambda x: f"€{x:,.2f}")
    df_mostrar["Saldo Teórico"] = df_mostrar["Saldo Teórico"].apply(lambda x: f"€{x:,.2f}")
    df_mostrar[""] = df_mostrar["es_real"].apply(
        lambda x: "✅ Real" if x else "🔮 Proyectado")

    st.dataframe(
        df_mostrar[["Mes", "Balance Real", "Saldo Real", "Presupuesto", "Saldo Teórico", ""]],
        use_container_width=True, hide_index=True
    )

# ── Helpers de parseo del xlsx ────────────────────────────────────────────────

def _extraer_nombre_xlsx(val):
    if val is None:
        return None
    s = str(val).strip()
    if s.startswith("="):
        m = re.search(r'"([^"]+)"\s*\)$', s)
        return m.group(1) if m else None
    return s

def _extraer_precio_xlsx(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val)
    if s.startswith("="):
        m = re.search(r",\s*([\d.]+)\s*\)", s)
        return float(m.group(1)) if m else None
    return None

def procesar_xlsx_cartera(archivo):
    wb = openpyxl.load_workbook(archivo)
    if "INV Esp" not in wb.sheetnames:
        raise ValueError("No se encontró la pestaña 'INV Esp' en el archivo.")
    ws = wb["INV Esp"]

    sector_map = {}
    for row in ws.iter_rows(min_row=13, max_row=27, values_only=True):
        nombre, sector = row[0], row[1]
        if nombre and sector and not str(nombre).startswith("="):
            sector_map[nombre.strip()] = sector.strip()

    transacciones = []
    for row in ws.iter_rows(min_row=31, max_row=ws.max_row, values_only=True):
        if not any(v is not None for v in row[:6]):
            break
        nombre = _extraer_nombre_xlsx(row[0])
        ticker = row[1]
        fecha = row[2]
        tipo = row[3]
        cantidad = row[4]
        precio_entrada = row[5]
        precio_actual = _extraer_precio_xlsx(row[6])

        if not ticker or not fecha or not nombre:
            continue

        fecha_dt = fecha.date() if isinstance(fecha, datetime) else fecha
        sector = sector_map.get(nombre)

        transacciones.append({
            "ticker": str(ticker).strip(),
            "nombre": nombre,
            "sector": sector,
            "tipo": str(tipo).strip() if tipo else "Compra",
            "fecha_operacion": str(fecha_dt),
            "cantidad": float(cantidad) if cantidad else 0.0,
            "precio_entrada": float(precio_entrada) if precio_entrada else 0.0,
            "precio_actual": precio_actual,
            "moneda": "USD",
        })

    if not transacciones:
        raise ValueError("No se encontraron transacciones en la pestaña 'INV Esp'.")

    df_transacciones = pd.DataFrame(transacciones)
    df_tickers = (
        df_transacciones[["ticker", "nombre", "sector", "moneda"]]
        .drop_duplicates(subset="ticker")
        .reset_index(drop=True)
    )
    return df_transacciones, df_tickers

# ── Cartera: queries ──────────────────────────────────────────────────────────

@st.cache_data(ttl=300)
def get_carteras(_supabase, user_id):
    """Devuelve lista de carteras del usuario."""
    result = _supabase.table("carteras")\
        .select("id, nombre, moneda, created_at")\
        .eq("user_id", user_id)\
        .order("created_at")\
        .execute()
    return result.data or []

def crear_cartera(supabase, user_id, nombre, moneda):
    """Crea una nueva cartera. Devuelve el id generado."""
    result = supabase.table("carteras").insert({
        "user_id": user_id,
        "nombre": nombre,
        "moneda": moneda
    }).execute()
    return result.data[0]["id"] if result.data else None

@st.cache_data(ttl=300)
def get_cartera(_supabase, user_id, cartera_id):
    """
    Devuelve DataFrame con transacciones activas de una cartera específica
    + sector/nombre de cartera_tickers.
    """
    result_c = _supabase.table("cartera")\
        .select("id, ticker, tipo, fecha_operacion, cantidad, precio_entrada, precio_actual, comision")\
        .eq("user_id", user_id)\
        .eq("cartera_id", cartera_id)\
        .eq("estado", "activo")\
        .execute()

    result_t = _supabase.table("cartera_tickers")\
        .select("ticker, nombre, sector, moneda")\
        .eq("user_id", user_id)\
        .execute()

    if not result_c.data:
        return pd.DataFrame()

    df = pd.DataFrame(result_c.data)
    df["fecha_operacion"] = pd.to_datetime(df["fecha_operacion"])
    df["cantidad"] = df["cantidad"].astype(float)
    df["precio_entrada"] = df["precio_entrada"].astype(float)
    df["precio_actual"] = df["precio_actual"].astype(float)
    df["comision"] = df["comision"].fillna(0).astype(float)
    df["posicion_inicial"] = df["cantidad"] * df["precio_entrada"] + df["comision"]
    df["posicion_actual"] = df["cantidad"] * df["precio_actual"]
    df["gp"] = df["posicion_actual"] - df["posicion_inicial"]

    if result_t.data:
        df_tickers = pd.DataFrame(result_t.data)
        df = df.merge(df_tickers, on="ticker", how="left")

    return df

@st.cache_data(ttl=300)
def get_cartera_eliminada(_supabase, user_id, cartera_id):
    """Devuelve posiciones eliminadas (papelera) de una cartera."""
    result_c = _supabase.table("cartera")\
        .select("id, ticker, tipo, fecha_operacion, cantidad, precio_entrada, comision")\
        .eq("user_id", user_id)\
        .eq("cartera_id", cartera_id)\
        .eq("estado", "eliminado")\
        .execute()

    if not result_c.data:
        return pd.DataFrame()

    result_t = _supabase.table("cartera_tickers")\
        .select("ticker, nombre")\
        .eq("user_id", user_id)\
        .execute()

    df = pd.DataFrame(result_c.data)
    if result_t.data:
        df_tickers = pd.DataFrame(result_t.data)
        df = df.merge(df_tickers, on="ticker", how="left")

    return df

@st.cache_data(ttl=300)
def get_tickers_sin_sector(_supabase, user_id):
    result = _supabase.table("cartera_tickers")\
        .select("ticker, nombre")\
        .eq("user_id", user_id)\
        .is_("sector", "null")\
        .execute()
    return result.data or []

@st.cache_data(ttl=300)
def get_efectivo_actual(_supabase, user_id, cartera_id):
    """Devuelve el último snapshot de efectivo para una cartera."""
    result = _supabase.table("cartera_efectivo")\
        .select("monto, fecha_registro")\
        .eq("user_id", user_id)\
        .eq("cartera_id", cartera_id)\
        .order("fecha_registro", desc=True)\
        .limit(1)\
        .execute()
    if not result.data:
        return 0.0, None
    return float(result.data[0]["monto"]), result.data[0]["fecha_registro"]

def guardar_efectivo(supabase, user_id, cartera_id, monto):
    """Inserta un nuevo snapshot de efectivo (no sobrescribe historial)."""
    hoy = str(date.today())
    supabase.table("cartera_efectivo").insert({
        "user_id": user_id,
        "cartera_id": cartera_id,
        "monto": float(monto),
        "fecha_registro": hoy
    }).execute()

# ── Cartera: sincronización xlsx (carga inicial) ──────────────────────────────

def sincronizar_cartera(df_transacciones, df_tickers, supabase, user_id, cartera_id):
    """
    Carga inicial desde xlsx para una cartera específica.
    Borra solo las posiciones de esa cartera (no afecta otras).
    """
    if "access_token" in st.session_state and st.session_state.access_token:
        supabase.postgrest.auth(st.session_state.access_token)

    # 1. Upsert tickers
    for _, row in df_tickers.iterrows():
        supabase.table("cartera_tickers").upsert(
            {"ticker": row["ticker"], "nombre": row["nombre"],
             "sector": row["sector"], "moneda": row["moneda"], "user_id": user_id},
            on_conflict="ticker,user_id",
        ).execute()

    # 2. Borrar solo posiciones de esta cartera
    supabase.table("cartera")\
        .delete()\
        .eq("user_id", user_id)\
        .eq("cartera_id", cartera_id)\
        .execute()

    # 3. INSERT con cartera_id
    registros = df_transacciones.to_dict(orient="records")
    for r in registros:
        r["user_id"] = user_id
        r["cartera_id"] = cartera_id
        r["estado"] = "activo"
        r["comision"] = r.get("comision", 0) or 0
        r.pop("sector", None)
        r.pop("moneda", None)
        r.pop("nombre", None)

    for i in range(0, len(registros), 500):
        supabase.table("cartera").insert(registros[i:i + 500]).execute()

    return len(registros)

# ── Cartera: entrada manual ───────────────────────────────────────────────────

def formulario_nueva_posicion(supabase, user_id, cartera_id):
    """Formulario para añadir una posición manualmente."""
    with st.expander("➕ Añadir posición", expanded=False):
        with st.form(key=f"form_posicion_{cartera_id}"):
            col1, col2 = st.columns(2)
            with col1:
                ticker_input = st.text_input("Ticker", placeholder="Ej: AAPL, IDUS, BRK.B").upper().strip()
                mercado = st.selectbox("Mercado", [
                    "EE.UU. (NASDAQ/NYSE)",
                    "Londres (LSE)",
                    "Alemania (Xetra)",
                    "Francia (Euronext París)",
                    "Países Bajos (Euronext Ámsterdam)",
                    "Italia (Borsa Italiana)",
                    "España (BME)",
                    "Suiza (SIX)",
                    "Japón (Tokio)",
                    "Hong Kong",
                    "Canadá (Toronto)",
                ])
                tipo = st.selectbox("Tipo", ["Compra", "Venta"])
                fecha = st.date_input("Fecha de operación", value=date.today())
            with col2:
                cantidad = st.number_input("Cantidad", min_value=0.0, step=0.0001, format="%.4f")
                precio = st.number_input("Precio (entrada o salida)", min_value=0.0, step=0.0001, format="%.4f")
                comision = st.number_input("Comisión (opcional)", min_value=0.0, step=0.01, format="%.2f", value=0.0)

            submitted = st.form_submit_button("💾 Guardar posición", type="primary")

            if submitted:
                prefijo = MERCADO_A_PREFIJO[mercado]
                if prefijo and ticker_input and not ticker_input.startswith(prefijo):
                    ticker = f"{prefijo}{ticker_input}"
                else:
                    ticker = ticker_input
                if not ticker_input:
                    st.error("El ticker es obligatorio.")
                elif cantidad <= 0:
                    st.error("La cantidad debe ser mayor que 0.")
                elif precio <= 0:
                    st.error("El precio debe ser mayor que 0.")
                else:
                    try:
                        # Insertar en cartera
                        supabase.table("cartera").insert({
                            "user_id": user_id,
                            "cartera_id": cartera_id,
                            "ticker": ticker,
                            "tipo": tipo,
                            "fecha_operacion": str(fecha),
                            "cantidad": float(cantidad),
                            "precio_entrada": float(precio),
                            "precio_actual": float(precio),  # precio inicial = precio entrada
                            "comision": float(comision),
                            "estado": "activo"
                        }).execute()

                        # Obtener nombre real vía yfinance
                        nombre_real = ticker
                        try:
                            ticker_yf = convertir_ticker_yfinance(ticker)
                            if ticker_yf:
                                info = yf.Ticker(ticker_yf).info
                                nombre_real = info.get("longName") or info.get("shortName") or ticker
                        except Exception:
                            pass

                        # Upsert en cartera_tickers (no pisa nombre si ya existe uno bueno)
                        existente = supabase.table("cartera_tickers")\
                            .select("nombre").eq("ticker", ticker).eq("user_id", user_id).execute()
                        if existente.data and existente.data[0]["nombre"] != ticker:
                            nombre_real = existente.data[0]["nombre"]

                        supabase.table("cartera_tickers").upsert(
                            {"ticker": ticker, "nombre": nombre_real,
                             "user_id": user_id, "moneda": "USD"},
                            on_conflict="ticker,user_id"
                        ).execute()

                        get_cartera.clear()
                        get_tickers_sin_sector.clear()
                        st.success(f"✅ Posición {tipo} de {ticker} añadida correctamente")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error al guardar: {e}")


# ── Widget editable ───────────────────────────────────────────────────────────

def widget_efectivo(supabase, user_id, cartera_id, moneda_sym):
    """Widget para ver/actualizar el efectivo disponible en la cartera."""
    monto_actual, fecha = get_efectivo_actual(supabase, user_id, cartera_id)

    with st.expander(f"💵 Efectivo disponible: {moneda_sym}{monto_actual:,.2f}", expanded=False):
        if fecha:
            st.caption(f"Última actualización: {fecha}")
        nuevo_monto = st.number_input(
            "Actualizar efectivo disponible",
            min_value=0.0, step=0.01, value=monto_actual,
            format="%.2f", key=f"efectivo_{cartera_id}"
        )
        if st.button("💾 Guardar efectivo", key=f"btn_efectivo_{cartera_id}"):
            guardar_efectivo(supabase, user_id, cartera_id, nuevo_monto)
            get_efectivo_actual.clear()
            st.success("✅ Efectivo actualizado")
            st.rerun()

    return monto_actual


# ── Cartera: widget asignar sector ────────────────────────────────────────────

def widget_asignar_sector(supabase, user_id):
    tickers_sin_sector = get_tickers_sin_sector(supabase, user_id)
    if not tickers_sin_sector:
        return

    result = supabase.table("cartera_tickers")\
        .select("sector").eq("user_id", user_id).execute()
    sectores_existentes = sorted(
        set(r["sector"] for r in (result.data or []) if r.get("sector"))
    )

    st.warning(
        f"⚠️ {len(tickers_sin_sector)} ticker(s) sin sector asignado. "
        "Asígnalos para que aparezcan en los resúmenes."
    )

    for item in tickers_sin_sector:
        ticker = item["ticker"]
        nombre = item.get("nombre", ticker)
        col1, col2, col3 = st.columns([2, 2, 1])
        with col1:
            st.markdown(f"**{ticker}** — {nombre}")
        with col2:
            opciones = sectores_existentes + ["➕ Nuevo sector..."]
            seleccion = st.selectbox("Sector", opciones,
                                     key=f"sector_sel_{ticker}", label_visibility="collapsed")
            if seleccion == "➕ Nuevo sector...":
                seleccion = st.text_input("Nuevo sector", key=f"sector_nuevo_{ticker}",
                                          placeholder="Ej: Healthcare")
        with col3:
            if st.button("💾 Guardar", key=f"sector_btn_{ticker}"):
                if seleccion and seleccion != "➕ Nuevo sector...":
                    supabase.table("cartera_tickers").update(
                        {"sector": seleccion}
                    ).eq("ticker", ticker).eq("user_id", user_id).execute()
                    get_cartera.clear()
                    get_tickers_sin_sector.clear()
                    st.success(f"✅ {ticker} → {seleccion}")
                    st.rerun()

# ── Cartera: conversión tickers yfinance ──────────────────────────────────────
MERCADO_A_PREFIJO = {
    "EE.UU. (NASDAQ/NYSE)": "",
    "Londres (LSE)": "LON:",
    "Alemania (Xetra)": "GER:",
    "Francia (Euronext París)": "PAR:",
    "Países Bajos (Euronext Ámsterdam)": "AMS:",
    "Italia (Borsa Italiana)": "MIL:",
    "España (BME)": "MAD:",
    "Suiza (SIX)": "SIX:",
    "Japón (Tokio)": "TYO:",
    "Hong Kong": "HKG:",
    "Canadá (Toronto)": "TOR:",
}


PREFIJO_A_SUFIJO_YF = {
    "LON:": ".L",    # Londres (LSE)
    "GER:": ".DE",   # Alemania (Xetra)
    "PAR:": ".PA",   # París (Euronext)
    "AMS:": ".AS",   # Ámsterdam (Euronext)
    "MIL:": ".MI",   # Milán (Borsa Italiana)
    "MAD:": ".MC",   # Madrid (BME)
    "SIX:": ".SW",   # Suiza (SIX)
    "TYO:": ".T",    # Tokio
    "HKG:": ".HK",   # Hong Kong
    "TOR:": ".TO",   # Toronto
}

def convertir_ticker_yfinance(ticker_original):
    """
    Convierte el ticker guardado al formato que acepta yfinance.

    Reglas:
      PREFIJO:XXX → XXX.suf  (según PREFIJO_A_SUFIJO_YF)
      BRK.B    → BRK-B      (Berkshire: yfinance usa guión)
      GLDV, CSPX, IGLN → XXX.L  (ETFs europeos legacy sin prefijo)
      Cash / Efectivo → None  (ignorar, no es un ticker)
      Resto → sin cambio     (NYSE/NASDAQ estándar)
    """
    if not ticker_original:
        return None
    t = str(ticker_original).strip()
    if t.lower() in ("cash", "efectivo", ""):
        return None

    t_upper = t.upper()
    for prefijo, sufijo in PREFIJO_A_SUFIJO_YF.items():
        if t_upper.startswith(prefijo):
            return t_upper[len(prefijo):] + sufijo

    if t_upper == "BRK.B":
        return "BRK-B"

    ETFS_LONDON = {"GLDV", "CSPX", "IGLN"}
    if t_upper in ETFS_LONDON:
        return t_upper + ".L"

    return t

@st.cache_data(ttl=3600)
def obtener_tipo_cambio_usd_eur():
    """Devuelve cuántos EUR equivale 1 USD. Cacheado 1h."""
    try:
        data = yf.download("EURUSD=X", period="1d", interval="1d",
                            auto_adjust=True, progress=False)
        eur_usd = float(data["Close"].iloc[-1])  # cuántos USD por 1 EUR
        return 1 / eur_usd  # cuántos EUR por 1 USD
    except Exception:
        return None  # fallback: el prompt indica "no disponible"

def get_precios_yfinance(tickers_originales):
    precios = {}
    errores = {}
    mapa = {}
    for t_orig in tickers_originales:
        t_yf = convertir_ticker_yfinance(t_orig)
        if t_yf:
            mapa[t_yf] = t_orig

    if not mapa:
        return precios, errores

    try:
        tickers_yf = list(mapa.keys())
        data = yf.download(tickers=tickers_yf, period="1d", interval="1d",
                           auto_adjust=True, progress=False, threads=True)
        if not data.empty:
            close = data["Close"] if "Close" in data.columns else data
            for t_yf, t_orig in mapa.items():
                try:
                    serie = close if len(tickers_yf) == 1 else close[t_yf]
                    serie_limpia = serie.dropna()
                    if not serie_limpia.empty:
                        precios[t_orig] = round(float(serie_limpia.iloc[-1]), 4)
                    else:
                        errores[t_orig] = "Sin datos de precio"
                except Exception:
                    errores[t_orig] = f"Ticker no reconocido: {t_yf}"
        else:
            for t_orig in mapa.values():
                errores[t_orig] = "Sin datos (posible fallo de conexión)"
    except Exception:
        for t_orig in mapa.values():
            errores[t_orig] = "yfinance no disponible"

    return precios, errores

# ── Cartera: vistas internas ──────────────────────────────────────────────────

def color_gp(val):
    try:
        num = float(val.replace('$','').replace(',','').replace('%',''))
        return 'color: #2ecc71' if num >= 0 else 'color: #e74c3c'
    except Exception:
        return ''

def _vista_por_sector(df):
    if "sector" not in df.columns:
        st.warning("Sin datos de sector.")
        return

    # Calcular posición neta por ticker (compras - ventas)
    df_compras = df[df["tipo"] == "Compra"].groupby("ticker").agg(
        cantidad_c=("cantidad", "sum"),
        valor_c=("posicion_inicial", "sum"),
        valor_actual_c=("posicion_actual", "sum"),
    ).reset_index()

    df_ventas = df[df["tipo"] == "Venta"].groupby("ticker").agg(
        cantidad_v=("cantidad", "sum"),
        valor_v=("posicion_inicial", "sum"),
        valor_actual_v=("posicion_actual", "sum"),
    ).reset_index()

    df_neto = df_compras.merge(df_ventas, on="ticker", how="left")
    df_neto = df_neto.fillna(0)
    df_neto["cantidad_neta"] = df_neto["cantidad_c"] - df_neto["cantidad_v"]
    df_neto["posicion_inicial"] = df_neto["valor_c"] - df_neto["valor_v"]
    df_neto["posicion_actual"] = df_neto["valor_actual_c"] - df_neto["valor_actual_v"]

    # Excluir posiciones cerradas (cantidad neta = 0)
    df_neto = df_neto[df_neto["cantidad_neta"] > 0.0001]

    # Añadir sector
    df_sector_map = df[["ticker", "sector"]].drop_duplicates("ticker")
    df_neto = df_neto.merge(df_sector_map, on="ticker", how="left")

    resumen = df_neto.groupby("sector").agg(
        posicion_inicial=("posicion_inicial", "sum"),
        posicion_actual=("posicion_actual", "sum"),
    ).reset_index()
    resumen["gp"] = resumen["posicion_actual"] - resumen["posicion_inicial"]
    resumen["gp_pct"] = resumen["gp"] / resumen["posicion_inicial"] * 100
    total_actual = resumen["posicion_actual"].sum()
    resumen["pct_total"] = resumen["posicion_actual"] / total_actual * 100
    resumen = resumen.sort_values("posicion_actual", ascending=False)

    df_mostrar = resumen.copy()
    df_mostrar["posicion_actual"] = df_mostrar["posicion_actual"].apply(lambda x: f"${x:,.2f}")
    df_mostrar["gp"] = df_mostrar["gp"].apply(lambda x: f"${x:,.2f}")
    df_mostrar["gp_pct"] = df_mostrar["gp_pct"].apply(lambda x: f"{x:.2f}%")
    df_mostrar["pct_total"] = df_mostrar["pct_total"].apply(lambda x: f"{x:.2f}%")
    df_mostrar.columns = ["Sector", "Invertido", "Valor Actual", "G/P", "G/P %", "% Total"]
    df_mostrar = df_mostrar[["Sector", "Valor Actual", "% Total", "G/P", "G/P %"]]
    styled = df_mostrar.style.map(color_gp, subset=["G/P", "G/P %"])
    st.dataframe(styled, use_container_width=True, hide_index=True)

    st.divider()
    col_izq, col_der = st.columns(2)
    with col_izq:
        fig = px.pie(resumen, values="posicion_actual", names="sector", hole=0.45,
                     title="Distribución por sector",
                     color_discrete_sequence=px.colors.qualitative.Set2)
        fig.update_traces(textposition="inside", textinfo="percent+label")
        fig.update_layout(showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
    with col_der:
        st.markdown("#### 📈 Evolución histórica")
        st.info("📊 Próximamente — valor de mercado vs coste a lo largo del tiempo")

def _vista_por_ticker(df, precios_rt=None, errores_rt=None):
    # Calcular posición neta por ticker (compras - ventas)
    df_compras = df[df["tipo"] == "Compra"].groupby(["ticker", "nombre", "sector"]).apply(
        lambda g: pd.Series({
            "cantidad_c": g["cantidad"].sum(),
            "valor_c": (g["precio_entrada"] * g["cantidad"]).sum(),
            "precio_actual_xlsx": g.sort_values("fecha_operacion").iloc[-1]["precio_actual"],
        })
    ).reset_index()

    df_ventas = df[df["tipo"] == "Venta"].groupby("ticker").agg(
        cantidad_v=("cantidad", "sum"),
        valor_v=("posicion_inicial", "sum"),
    ).reset_index()

    resumen = df_compras.merge(df_ventas, on="ticker", how="left")
    resumen = resumen.fillna(0)
    resumen["cantidad"] = resumen["cantidad_c"] - resumen["cantidad_v"]
    resumen["precio_medio"] = resumen["valor_c"] / resumen["cantidad_c"]

    # Excluir posiciones cerradas
    resumen = resumen[resumen["cantidad"] > 0.0001]

    def resolver_precio(row):
        if precios_rt and row["ticker"] in precios_rt:
            return precios_rt[row["ticker"]]
        return row["precio_actual_xlsx"]

    resumen["precio_final"] = resumen.apply(resolver_precio, axis=1)
    resumen["posicion_inicial"] = resumen["cantidad"] * resumen["precio_medio"]
    resumen["posicion_actual"] = resumen["cantidad"] * resumen["precio_final"]
    resumen["gp"] = resumen["posicion_actual"] - resumen["posicion_inicial"]
    total = resumen["posicion_actual"].sum()
    resumen["pct_total"] = resumen["posicion_actual"] / total * 100
    resumen = resumen.sort_values("posicion_actual", ascending=False)

    df_mostrar = resumen.copy()
    df_mostrar["cantidad"] = df_mostrar["cantidad"].apply(lambda x: f"{x:.4f}")
    df_mostrar["posicion_actual"] = df_mostrar["posicion_actual"].apply(lambda x: f"${x:,.2f}")
    df_mostrar["gp"] = df_mostrar["gp"].apply(lambda x: f"${x:,.2f}")
    df_mostrar["pct_total"] = df_mostrar["pct_total"].apply(lambda x: f"{x:.2f}%")

    def fuente_precio(ticker):
        if precios_rt and ticker in precios_rt:
            return "🟢 RT"
        if errores_rt and ticker in errores_rt:
            return "⚠️ xlsx"
        return "📄 xlsx"

    df_mostrar["Precio"] = resumen["ticker"].apply(fuente_precio)
    df_mostrar = df_mostrar[["nombre", "sector", "cantidad", "posicion_actual", "gp", "pct_total", "Precio"]]
    df_mostrar.columns = ["Activo", "Sector", "Cantidad", "Posición", "G/P", "% Total", "Precio"]
    styled = df_mostrar.style.map(color_gp, subset=["G/P", "% Total"])
    st.dataframe(styled, use_container_width=True, hide_index=True)

    if errores_rt:
        tickers_fallback = [t for t in errores_rt if t != "_global"]
        if tickers_fallback:
            st.caption(f"⚠️ Precio xlsx (no RT): {', '.join(tickers_fallback)}.")

    st.divider()
    col_izq, col_der = st.columns(2)
    with col_izq:
        fig = px.pie(resumen, values="posicion_actual", names="nombre", hole=0.45,
                     title="Distribución por activo",
                     color_discrete_sequence=px.colors.qualitative.Set2)
        fig.update_traces(textposition="inside", textinfo="percent+label")
        fig.update_layout(showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
    with col_der:
        st.markdown("#### 📈 Evolución histórica")
        st.info("📊 Próximamente — valor de mercado vs coste a lo largo del tiempo")

def _vista_historial(df, supabase, user_id, cartera_id):
    """Historial en tabla con botón de soft delete por posición."""
    mostrar_papelera = st.toggle("🗑️ Ver posiciones eliminadas", value=False,
                                  key=f"papelera_{cartera_id}")

    if mostrar_papelera:
        df_elim = get_cartera_eliminada(supabase, user_id, cartera_id)
        if df_elim.empty:
            st.info("No hay posiciones eliminadas.")
        else:
            st.subheader("🗑️ Posiciones eliminadas")
            for _, row in df_elim.iterrows():
                col1, col2 = st.columns([5, 1])
                with col1:
                    nombre = row.get("nombre", row["ticker"])
                    st.markdown(
                        f"**{nombre}** ({row['ticker']}) · {row['tipo']} · "
                        f"{pd.to_datetime(row['fecha_operacion']).strftime('%d/%m/%Y')} · "
                        f"{float(row['cantidad']):.4f} × ${float(row['precio_entrada']):,.3f}"
                    )
                with col2:
                    if st.button("↩️ Restaurar", key=f"restaurar_{row['id']}"):
                        supabase.table("cartera").update({"estado": "activo"})\
                            .eq("id", row["id"]).execute()
                        get_cartera.clear()
                        get_cartera_eliminada.clear()
                        st.success("✅ Posición restaurada")
                        st.rerun()
        return

    # Vista tabla normal
    df_hist = df.sort_values("fecha_operacion", ascending=False).copy()
    df_hist["Fecha"] = df_hist["fecha_operacion"].dt.strftime("%d/%m/%Y")
    df_hist["Activo"] = df_hist.get("nombre", df_hist["ticker"])
    df_hist["Cantidad"] = df_hist["cantidad"].apply(lambda x: f"{x:.4f}")
    df_hist["Precio"] = df_hist["precio_entrada"].apply(lambda x: f"${x:,.3f}")
    df_hist["Precio actual"] = df_hist["precio_actual"].apply(lambda x: f"${x:,.3f}")
    df_hist["Posición inicial"] = df_hist["posicion_inicial"].apply(lambda x: f"${x:,.2f}")
    df_hist["Posición actual"] = df_hist["posicion_actual"].apply(lambda x: f"${x:,.2f}")
    df_hist["G/P"] = df_hist["gp"].apply(lambda x: f"${x:,.2f}")

    cols_mostrar = ["Fecha", "Activo", "ticker", "tipo", "Cantidad",
                    "Precio", "Precio actual", "Posición inicial", "Posición actual", "G/P"]
    df_tabla = df_hist[cols_mostrar].rename(columns={"ticker": "Ticker", "tipo": "Tipo"})

    styled = df_tabla.style.map(color_gp, subset=["G/P"])
    st.dataframe(styled, use_container_width=True, hide_index=True)

    # Botones de eliminar debajo de la tabla
    st.caption("Selecciona una posición para eliminarla:")
    for _, row in df_hist.iterrows():
        col1, col2 = st.columns([5, 1])
        with col1:
            nombre = row.get("Activo", row["ticker"])
            st.caption(
                f"{row['Fecha']} · {nombre} · {row['tipo']} · "
                f"{row['Cantidad']} × {row['Precio']}"
            )
        with col2:
            if st.button("🗑️", key=f"del_pos_{row['id']}", help="Eliminar"):
                supabase.table("cartera").update({"estado": "eliminado"})\
                    .eq("id", row["id"]).execute()
                get_cartera.clear()
                st.info("Posición eliminada. Recupérala con el toggle de arriba.")
                st.rerun()

# ── Master Prompt Engine: helpers ─────────────────────────────────────────────
# Creación de Tabla Mardown para el Master Prompt

def generar_tabla_anual(supabase, user_id, year):
    """Tabla anual: Presupuesto Anual vs Real YTD por categoría + balance YTD.
    Si no hay presupuestos cargados, muestra solo gastos/ingresos reales YTD."""
    hoy = date.today()

    # % del año transcurrido (usado para la columna "Desvío")
    primer_dia_anio = date(year, 1, 1)
    primer_dia_anio_siguiente = date(year + 1, 1, 1)
    dias_en_anio = (primer_dia_anio_siguiente - primer_dia_anio).days
    dia_del_anio = (hoy - primer_dia_anio).days + 1
    pct_anio_transcurrido = dia_del_anio / dias_en_anio * 100

    df_todos = get_todos_gastos(supabase, user_id)
    if not df_todos.empty:
        df_ytd = df_todos[
            (df_todos["anio"] == year) &
            (df_todos["fecha_gasto"].dt.date <= hoy)
        ]
        real_ytd = df_ytd.groupby("categoria_consumo")["importe"].sum().reset_index()
        real_ytd.columns = ["categoria_consumo", "real_ytd"]
        balance_ytd = real_ytd["real_ytd"].sum()
    else:
        real_ytd = pd.DataFrame(columns=["categoria_consumo", "real_ytd"])
        balance_ytd = 0.0

    df_presup_anio = get_presupuestos_anio(supabase, year, user_id)
    mask_otros_real = real_ytd["categoria_consumo"].isin(CATEGORIAS_OTROS)

    if df_presup_anio.empty:
        # ── Sin presupuesto (ej. Alicia): solo reales YTD ──
        df_principales = real_ytd[~mask_otros_real].copy()
        df_otros = real_ytd[mask_otros_real].copy()
        df_principales.columns = ["Categoría", "Real YTD"]
        df_principales = df_principales.reindex(
            df_principales["Real YTD"].abs().sort_values(ascending=False).index
        )

        lineas = ["| Categoría | Real YTD |", "|---|---|"]
        for _, row in df_principales.iterrows():
            lineas.append(f"| {row['Categoría']} | €{row['Real YTD']:,.2f} |")

        nota = ""
        if not df_otros.empty:
            otros_sum = df_otros["real_ytd"].sum()
            lineas.append(f"| Otras Categorías | €{otros_sum:,.2f} |")
            cats_lista = ", ".join(sorted(df_otros["categoria_consumo"].tolist()))
            nota = f"\n\n_Otras Categorías incluye: {cats_lista}._"

        nota += "\n\n_Sin presupuesto configurado — solo gastos/ingresos reales acumulados en el año (YTD)._"
        return "\n".join(lineas) + nota, balance_ytd

    # ── Con presupuesto: Presupuesto Anual | Real YTD | % Alcance ──
    df_tabla = pd.merge(df_presup_anio, real_ytd, on="categoria_consumo", how="outer").fillna(0)
    df_tabla.columns = ["Categoría", "Presupuesto Anual", "Real YTD"]

    mask_otros = df_tabla["Categoría"].isin(CATEGORIAS_OTROS)
    df_principales = df_tabla[~mask_otros].copy()
    df_otros = df_tabla[mask_otros].copy()

    if not df_otros.empty:
        fila_otros = pd.DataFrame([{
            "Categoría": "Otras Categorías",
            "Presupuesto Anual": df_otros["Presupuesto Anual"].sum(),
            "Real YTD": df_otros["Real YTD"].sum(),
        }])
        df_final = pd.concat([df_principales, fila_otros], ignore_index=True)
        cats_lista = ", ".join(sorted(df_otros["Categoría"].tolist()))
        nota = f"\n\n_Otras Categorías incluye: {cats_lista}._"
    else:
        df_final = df_principales
        nota = ""

    def calc_alcance(row):
        if row["Presupuesto Anual"] == 0:
            return None
        return row["Real YTD"] / row["Presupuesto Anual"] * 100

    df_final["% Alcance"] = df_final.apply(calc_alcance, axis=1)
    df_final = df_final.reindex(
        df_final["Real YTD"].abs().sort_values(ascending=False).index
    )

    lineas = ["| Categoría | Presupuesto Anual | Real YTD | % Alcance | Desvío |", "|---|---|---|---|---|"]
    for _, row in df_final.iterrows():
        if pd.notna(row["% Alcance"]):
            alcance = f"{row['% Alcance']:.0f}%"
            desvio = f"{row['% Alcance'] - pct_anio_transcurrido:+.0f} pp"
        else:
            alcance = "—"
            desvio = "—"
        lineas.append(
            f"| {row['Categoría']} | €{row['Presupuesto Anual']:,.2f} | "
            f"€{row['Real YTD']:,.2f} | {alcance} | {desvio} |"
        )

    nota += (
        f"\n\n_Desvío = % Alcance − % año transcurrido ({pct_anio_transcurrido:.0f}%). "
        f"Un valor positivo indica una categoría ejecutándose por encima del ritmo esperado para esta fecha; "
        f"negativo, por debajo._"
        f"\n\n_Un valor positivo en Real YTD dentro de una categoría de gasto puede indicar un reembolso o devolución, no un ingreso._"
    )
    return "\n".join(lineas) + nota, balance_ytd

# Crear una tabla de Ingresos y Gastos mensuales en el Prompt Engine, para ver la evolución de cada mes del año en curso (YTD).

def generar_tabla_mensual_ingresos_gastos(supabase, user_id, year, categorias_inversion=None):
    """Tabla compacta: Ingresos, Gastos (sin inversión) y Aportaciones a inversión por mes (YTD)."""
    categorias_inversion = categorias_inversion or []
    hoy = date.today()
    df_todos = get_todos_gastos(supabase, user_id)
    if df_todos.empty:
        return "_Sin datos._"

    df_ytd = df_todos[
        (df_todos["anio"] == year) &
        (df_todos["fecha_gasto"].dt.date <= hoy)
    ]
    if df_ytd.empty:
        return "_Sin datos._"

    meses_es = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

    df_ingresos = df_ytd[df_ytd["tipo"] == "Ingreso"].groupby("mes")["monto"].sum()
    df_gastos_base = df_ytd[
        (df_ytd["tipo"] == "Gasto") &
        (~df_ytd["categoria_consumo"].isin(categorias_inversion))
    ].groupby("mes")["monto"].sum()

    if categorias_inversion:
        df_inversion = df_ytd[
            (df_ytd["tipo"] == "Gasto") &
            (df_ytd["categoria_consumo"].isin(categorias_inversion))
        ].groupby("mes")["monto"].sum()
        lineas = ["| Mes | Ingresos | Gastos (sin inversión) | Aportaciones a inversión | Tasa de ahorro |", "|---|---|---|---|---|"]
        for mes in sorted(df_ytd["mes"].unique()):
            ing = df_ingresos.get(mes, 0.0)
            gas = df_gastos_base.get(mes, 0.0)
            inv = df_inversion.get(mes, 0.0)
            tasa = f"{(ing - gas - inv) / ing * 100:.0f}%" if ing > 0 else "—"
            lineas.append(
                f"| {meses_es[mes]} | €{ing:,.2f} | €{gas:,.2f} | €{inv:,.2f} | {tasa} |"
            )
    else:
        lineas = ["| Mes | Ingresos | Gastos | Tasa de ahorro |", "|---|---|---|---|"]
        for mes in sorted(df_ytd["mes"].unique()):
            ing = df_ingresos.get(mes, 0.0)
            gas = df_gastos_base.get(mes, 0.0)
            tasa = f"{(ing - gas) / ing * 100:.0f}%" if ing > 0 else "—"
            lineas.append(
                f"| {meses_es[mes]} | €{ing:,.2f} | €{gas:,.2f} | {tasa} |"
            )

    return "\n".join(lineas)

# Replica el cálculo de KPIs de pagina_cartera() por cada cartera, más distribución por sector y últimas 10 compras: 

def generar_seccion_cartera(supabase, user_id):
    """Genera el bloque markdown de la sección 5, una por cada cartera activa."""
    carteras = get_carteras(supabase, user_id)
    if not carteras:
        return "_Sin carteras registradas._"

    tipo_cambio = obtener_tipo_cambio_usd_eur()  # EUR por 1 USD, o None

    bloques = []
    for c in carteras:
        cartera_id = c["id"]
        nombre = c["nombre"]
        moneda = c["moneda"]
        moneda_sym = "€" if moneda == "EUR" else "$"

        df = get_cartera(supabase, user_id, cartera_id)
        efectivo, _ = get_efectivo_actual(supabase, user_id, cartera_id)

        if df.empty:
            bloques.append(
                f"**Cartera: {nombre}** ({moneda})\n"
                f"- Sin posiciones activas.\n"
                f"- Efectivo disponible: {moneda_sym}{efectivo:,.2f}"
            )
            continue

        tickers_unicos = [
            t for t in df["ticker"].unique()
            if str(t).strip().lower() not in ("cash", "efectivo", "")
        ]
        precios_rt, _ = get_precios_yfinance(tickers_unicos)

        df_kpi = df.copy()
        df_kpi["precio_final"] = df_kpi.apply(
            lambda row: precios_rt[row["ticker"]] if row["ticker"] in precios_rt else row["precio_actual"],
            axis=1
        )
        df_kpi["posicion_actual_rt"] = df_kpi["cantidad"] * df_kpi["precio_final"]

        df_compras = df_kpi[df_kpi["tipo"] == "Compra"]
        df_ventas = df_kpi[df_kpi["tipo"] == "Venta"]

        precio_medio_compra = (
            df_compras.groupby("ticker")
            .apply(lambda g: (g["precio_entrada"] * g["cantidad"]).sum() / g["cantidad"].sum())
        )

        coste_ventas = df_ventas.apply(
            lambda row: row["cantidad"] * precio_medio_compra.get(row["ticker"], row["precio_entrada"]),
            axis=1
        ).sum()

        total_invertido = df_compras["posicion_inicial"].sum() - coste_ventas
        total_actual = df_compras["posicion_actual_rt"].sum() - df_ventas["posicion_actual_rt"].sum()
        total_actual_con_efectivo = total_actual + efectivo
        total_gp = total_actual - total_invertido
        pct_gp = (total_gp / total_invertido * 100) if total_invertido != 0 else 0

        conversion = ""
        if moneda == "USD" and tipo_cambio:
            conversion = f" (≈ €{total_actual_con_efectivo * tipo_cambio:,.2f})"
        elif moneda == "USD":
            conversion = " (tipo de cambio no disponible)"

        # Distribución por sector (posiciones netas)
        df_neto_c = df_compras.groupby(["ticker", "sector"]).agg(
            cantidad_c=("cantidad", "sum"),
            valor_actual_c=("posicion_actual_rt", "sum")
        ).reset_index()
        df_neto_v = df_ventas.groupby("ticker").agg(
            cantidad_v=("cantidad", "sum"),
            valor_actual_v=("posicion_actual_rt", "sum")
        ).reset_index()
        df_neto = df_neto_c.merge(df_neto_v, on="ticker", how="left").fillna(0)
        df_neto["cantidad_neta"] = df_neto["cantidad_c"] - df_neto["cantidad_v"]
        df_neto["valor_neto"] = df_neto["valor_actual_c"] - df_neto["valor_actual_v"]
        df_neto = df_neto[df_neto["cantidad_neta"] > 0.0001]

        sector_resumen = df_neto.groupby("sector")["valor_neto"].sum().reset_index()
        sector_resumen["pct"] = (sector_resumen["valor_neto"] / total_actual * 100) if total_actual else 0
        sector_resumen = sector_resumen.sort_values("pct", ascending=False)

        sector_lineas = ["| Sector | % Cartera |", "|---|---|"]
        for _, row in sector_resumen.iterrows():
            sector_lineas.append(f"| {row['sector']} | {row['pct']:.1f}% |")

    # Posiciones actuales netas por ticker
        df_compras_ticker = df_compras.groupby("ticker").agg(
            nombre=("nombre", "first"),
            cantidad=("cantidad", "sum"),
            precio_medio=("precio_entrada", lambda x: (x * df_compras.loc[x.index, "cantidad"]).sum() / df_compras.loc[x.index, "cantidad"].sum()),
            valor_actual=("posicion_actual_rt", "sum")
        ).reset_index()
        df_ventas_ticker = df_ventas.groupby("ticker").agg(
            cantidad_v=("cantidad", "sum"),
            valor_actual_v=("posicion_actual_rt", "sum")
        ).reset_index()
        df_posiciones = df_compras_ticker.merge(df_ventas_ticker, on="ticker", how="left").fillna(0)
        df_posiciones["cantidad_neta"] = df_posiciones["cantidad"] - df_posiciones["cantidad_v"]
        df_posiciones["valor_neto"] = df_posiciones["valor_actual"] - df_posiciones["valor_actual_v"]
        df_posiciones = df_posiciones[df_posiciones["cantidad_neta"] > 0.0001]
        df_posiciones["precio_actual_rt"] = df_posiciones["ticker"].map(
            lambda t: precios_rt.get(t, None)
        )
        df_posiciones["pct_cartera"] = (df_posiciones["valor_neto"] / total_actual_con_efectivo * 100) if total_actual_con_efectivo else 0
        df_posiciones = df_posiciones.sort_values("valor_neto", ascending=False)

        posiciones_lineas = ["| Activo | Precio medio | Precio actual | Valor | % Cartera |", "|---|---|---|---|---|"]
        for _, row in df_posiciones.iterrows():
            nombre_activo = row["nombre"] if row["nombre"] else row["ticker"]
            precio_actual_str = f"{moneda_sym}{row['precio_actual_rt']:,.2f}" if row["precio_actual_rt"] else "—"
            posiciones_lineas.append(
                f"| {nombre_activo} | {moneda_sym}{row['precio_medio']:,.2f} | "
                f"{precio_actual_str} | {moneda_sym}{row['valor_neto']:,.2f} | {row['pct_cartera']:.1f}% |"
            )

        bloque = (
            f"**Cartera: {nombre}** ({moneda})\n"
            f"- Efectivo disponible: {moneda_sym}{efectivo:,.2f}\n"
            f"- Valor total (posiciones + efectivo): {moneda_sym}{total_actual_con_efectivo:,.2f}{conversion}\n"
            f"- Invertido: {moneda_sym}{total_invertido:,.2f}\n"
            f"- G/P: {moneda_sym}{total_gp:,.2f} ({pct_gp:.2f}%)\n\n"
            f"Distribución por sector:\n" + "\n".join(sector_lineas) + "\n\n"
            f"Posiciones actuales:\n" + "\n".join(posiciones_lineas)
        )
        bloques.append(bloque)

    return "\n\n".join(bloques)

# ensamblamos todo con la plantilla

def generar_prompt_master(supabase, user_id, pais, perfil, categorias_inversion=None,
                           deuda_importe=0, deuda_cuota=0, deuda_fecha_fin="", contexto_estrategico=""):
    """Ensambla el prompt completo para pegar en un LLM externo."""
    categorias_inversion = categorias_inversion or []
    hoy = date.today()
    year = hoy.year

    df_saldos, _ = get_saldos_actuales(supabase, user_id)
    total_bancos = df_saldos["monto"].sum() if not df_saldos.empty else 0.0

    tabla_anual, balance_ytd = generar_tabla_anual(supabase, user_id, year)
    tabla_mensual = generar_tabla_mensual_ingresos_gastos(supabase, user_id, year, categorias_inversion)

    # Meses de cobertura
    df_todos = get_todos_gastos(supabase, user_id)
    if not df_todos.empty:
        df_ytd = df_todos[
            (df_todos["anio"] == year) &
            (df_todos["fecha_gasto"].dt.date <= hoy) &
            (df_todos["tipo"] == "Gasto")
        ]
        meses_con_datos = df_ytd["mes"].nunique()
        gasto_total_ytd = df_ytd["monto"].sum()
        if meses_con_datos > 0 and gasto_total_ytd > 0:
            gasto_promedio_mensual = gasto_total_ytd / meses_con_datos
            meses_cobertura = total_bancos / gasto_promedio_mensual
            linea_cobertura = f"- Meses de cobertura (liquidez / gasto medio mensual): {meses_cobertura:.1f} meses"
        else:
            linea_cobertura = ""
    else:
        linea_cobertura = ""

    saldo_inicial_anio = total_bancos - balance_ytd

    primer_dia_anio = date(year, 1, 1)
    primer_dia_anio_siguiente = date(year + 1, 1, 1)
    dias_en_anio = (primer_dia_anio_siguiente - primer_dia_anio).days
    dia_del_anio = (hoy - primer_dia_anio).days + 1
    progreso_anio_pct = round(dia_del_anio / dias_en_anio * 100)

    seccion_cartera = generar_seccion_cartera(supabase, user_id)

    return f"""Actúa como mi asesor financiero personal. Voy a darte mi situación financiera completa. Tu trabajo es analizarla con honestidad, sin suavizar diagnósticos, y darme un plan de acción concreto.

### 1. CONTEXTO
- País: {pais}
- Perfil de inversión: {perfil}
- Moneda principal de gastos: €
- Fecha del informe: {hoy.strftime('%d/%m/%Y')}
{f"- Contexto estratégico: {contexto_estrategico}" if contexto_estrategico.strip() else ""}

### 2. SALDOS BANCARIOS
- Total líquido: €{total_bancos:,.2f}
{linea_cobertura}
{(f"- Deuda pendiente: €{deuda_importe:,.0f} — cuota €{deuda_cuota:,.0f}/mes hasta {deuda_fecha_fin}. El efectivo bancario cubre parcialmente esta obligación y no debe considerarse caja libre para inversión.") if deuda_importe > 0 else ""}

### 3. AÑO {year}
- Presupuesto Anual = suma de los 12 presupuestos mensuales. Real YTD = acumulado real desde el 1 de enero. % Alcance = Real YTD / Presupuesto Anual._

{tabla_anual}

- Saldo bancario al 1 de enero de {year} (estimado): €{saldo_inicial_anio:,.2f}
- Balance neto acumulado en {year} (YTD): €{balance_ytd:,.2f}
- Progreso del año: día {dia_del_anio} de {dias_en_anio} ({progreso_anio_pct}%)

**Evolución mensual (Ingresos vs Gastos):**

{tabla_mensual}

### 4. CARTERA DE INVERSIÓN

{seccion_cartera}

### INSTRUCCIONES
Mi perfil de inversión es {perfil} y vivo en {pais}. Con esa referencia, evalúa mis datos y responde en este orden:

1. **Diagnóstico general** (máx. 5 líneas): ¿cómo está mi salud financiera este año frente a mi plan?
2. **Gastos**: ¿qué categorías van adelantadas o atrasadas respecto al % del año transcurrido, y es motivo de preocupación o es normal?
3. **Cartera**: ¿mi diversificación y concentración son coherentes con mi perfil de riesgo declarado? Señala desbalances si los hay.
4. **Liquidez**: ¿tengo suficiente efectivo (banco + bróker, por moneda) para cubrir imprevistos o aprovechar oportunidades de inversión?
5. **Plan de acción**: 3 acciones concretas para los próximos 30 días, ordenadas de mayor a menor impacto.

No repitas los números que te doy — interprétalos. Si algo está mal, dilo sin rodeos."""

#  UI de la Pagina Prompt Engine

def pagina_prompt(supabase, user_id):
    st.title("🤖 Prompt Maestro")
    st.caption("Genera un resumen de tu situación financiera para analizar con ChatGPT, Gemini, Claude, etc.")

    col1, col2 = st.columns(2)
    with col1:
        pais = st.selectbox("País", ["España", "Otro"], index=0)
    with col2:
        perfil = st.selectbox(
            "Perfil de inversión",
            ["Conservador", "Moderado-conservador", "Moderado", "Moderado-agresivo", "Agresivo"],
            index=2
        )

    categorias_disponibles = get_categorias_usuario(supabase, user_id)
    default_inversion = [c for c in categorias_disponibles if "invest" in c.lower()]

    categorias_inversion = st.multiselect(
        "¿Qué categorías representan inversiones?",
        options=categorias_disponibles,
        default=default_inversion,
        help="Afecta cómo se interpreta 'Gastos' en la evolución mensual."
    )

    # Deudas (condicional)
    tiene_deudas = st.radio("¿Tienes deudas relevantes?", ["No", "Sí"], horizontal=True)
    if tiene_deudas == "Sí":
        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            deuda_importe = st.number_input("Importe total (€)", min_value=0, step=100, value=0)
        with col_d2:
            deuda_cuota = st.number_input("Cuota mensual (€)", min_value=0, step=50, value=0)
        with col_d3:
            deuda_fecha_fin = st.text_input("Fecha fin (MM/AAAA)", placeholder="12/2028")
    else:
        deuda_importe, deuda_cuota, deuda_fecha_fin = 0, 0, ""

    # Contexto estratégico guiado
    contexto_estrategico = st.text_area(
        "Contexto estratégico (opcional)",
        placeholder=(
            "Ej: Mi liquidez cubre una deuda — no es caja libre.\n"
            "XTB es mi bloque de crecimiento, no mi patrimonio total.\n"
            "Travel no tiene ejecuciones previstas hasta noviembre."
        ),
        max_chars=300,
        help="Se incluye en el prompt para que el LLM entienda tu estrategia global."
    )

    if st.button("🔄 Generar prompt", type="primary"):
        with st.spinner("Generando..."):
            try:
                st.session_state.prompt_generado = generar_prompt_master(
                    supabase, user_id, pais, perfil, categorias_inversion,
                    deuda_importe, deuda_cuota, deuda_fecha_fin, contexto_estrategico
                )
            except Exception as e:
                st.error(f"❌ Error al generar el prompt: {e}")

    if "prompt_generado" in st.session_state:
        st.divider()
        st.code(st.session_state.prompt_generado, language="markdown")
        st.caption("📋 Pasa el cursor sobre el bloque para copiarlo.")

# ── Página cartera ────────────────────────────────────────────────────────────

def pagina_cartera(supabase, user_id):
    st.title("💼 Cartera de Inversiones")

    widget_asignar_sector(supabase, user_id)

    # Cargar carteras del usuario
    carteras = get_carteras(supabase, user_id)

    # ── Formulario nueva cartera ──────────────────────────────────────────────
    LIMITE_CARTERAS = 5
    if len(carteras) < LIMITE_CARTERAS:
        with st.expander("➕ Nueva cartera"):
            with st.form("form_nueva_cartera"):
                col1, col2 = st.columns(2)
                with col1:
                    nombre_cartera = st.text_input("Nombre", placeholder="Ej: XTB Europa")
                with col2:
                    moneda_cartera = st.selectbox("Moneda", ["EUR", "USD", "GBP", "CHF"])
                crear = st.form_submit_button("Crear cartera", type="primary")
                if crear:
                    if not nombre_cartera.strip():
                        st.error("El nombre es obligatorio.")
                    else:
                        nuevo_id = crear_cartera(supabase, user_id,
                                                  nombre_cartera.strip(), moneda_cartera)
                        get_carteras.clear()
                        st.success(f"✅ Cartera '{nombre_cartera}' creada.")
                        st.rerun()
    else:
        st.caption(f"Has alcanzado el límite de {LIMITE_CARTERAS} carteras.")

    if not carteras:
        st.info("Aún no tienes carteras. Crea una para empezar.")
        return

    # ── Tabs por cartera ──────────────────────────────────────────────────────
    nombres_tabs = [f"{'💶' if c['moneda'] == 'EUR' else '💵'} {c['nombre']}" for c in carteras]
    tabs = st.tabs(nombres_tabs)

    for tab, cartera in zip(tabs, carteras):
        with tab:
            cartera_id = cartera["id"]
            moneda_sym = "€" if cartera["moneda"] == "EUR" else "$"

            # Formulario añadir posición
            formulario_nueva_posicion(supabase, user_id, cartera_id)
            efectivo_actual = widget_efectivo(supabase, user_id, cartera_id, moneda_sym)

            with st.spinner("Cargando cartera..."):
                df = get_cartera(supabase, user_id, cartera_id)

            if df.empty:
                st.info("No hay posiciones en esta cartera. Añade una manualmente o sube un xlsx.")
                continue

            # Precios en tiempo real
            tickers_unicos = [
                t for t in df["ticker"].unique()
                if str(t).strip().lower() not in ("cash", "efectivo", "")
            ]
            with st.spinner("Consultando precios en tiempo real..."):
                precios_rt, errores_rt = get_precios_yfinance(tickers_unicos)

            # Banner estado yfinance
            ahora = datetime.now().strftime("%H:%M")
            n_ok = len(precios_rt)
            if n_ok == len(tickers_unicos):
                st.success(f"🟢 Precios en tiempo real · Actualizado {ahora}")
            elif n_ok == 0:
                es_fallo_global = errores_rt and "yfinance no disponible" in list(errores_rt.values())[0]
                st.warning("🔴 yfinance no disponible · Usando precios del xlsx" if es_fallo_global
                           else "🔴 Sin precios en tiempo real · Usando precios del xlsx")
            else:
                st.warning(f"⚠️ {n_ok}/{len(tickers_unicos)} tickers con precio RT · Actualizado {ahora}")

            # KPIs
            df_kpi = df.copy()
            df_kpi["precio_final"] = df_kpi.apply(
                lambda row: precios_rt[row["ticker"]] if row["ticker"] in precios_rt else row["precio_actual"],
                axis=1
            )
            df_kpi["posicion_actual_rt"] = df_kpi["cantidad"] * df_kpi["precio_final"]
            df_kpi["gp_rt"] = df_kpi["posicion_actual_rt"] - df_kpi["posicion_inicial"]

# Invertido neto: para cada ticker vendido, restar el COSTE de compra
            # de las unidades vendidas (no el precio de venta)
            df_compras = df_kpi[df_kpi["tipo"] == "Compra"]
            df_ventas  = df_kpi[df_kpi["tipo"] == "Venta"]

            precio_medio_compra = (
                df_compras.groupby("ticker")
                .apply(lambda g: (g["precio_entrada"] * g["cantidad"]).sum() / g["cantidad"].sum())
            )

            coste_ventas = df_ventas.apply(
                lambda row: row["cantidad"] * precio_medio_compra.get(row["ticker"], row["precio_entrada"]),
                axis=1
            ).sum()

            total_invertido = df_compras["posicion_inicial"].sum() - coste_ventas
            total_actual    = df_compras["posicion_actual_rt"].sum() - df_ventas["posicion_actual_rt"].sum()
            total_actual_con_efectivo = total_actual + efectivo_actual
            total_gp        = total_actual - total_invertido
            pct_gp          = (total_gp / total_invertido * 100) if total_invertido != 0 else 0

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("💰 Invertido",    f"{moneda_sym}{total_invertido:,.2f}")
            k2.metric("📈 Valor actual", f"{moneda_sym}{total_actual_con_efectivo:,.2f}",
                     help=f"Incluye {moneda_sym}{efectivo_actual:,.2f} en efectivo")
            k3.metric("💹 G/P total",    f"{moneda_sym}{total_gp:,.2f}")
            k4.metric("📊 Rentabilidad", f"{pct_gp:.2f}%")

            st.divider()

            tab1, tab2, tab3 = st.tabs(["🗂️ Por Sector", "📋 Por Ticker", "📜 Historial"])
            with tab1:
                _vista_por_sector(df)
            with tab2:
                _vista_por_ticker(df, precios_rt=precios_rt, errores_rt=errores_rt)
            with tab3:
                _vista_historial(df, supabase, user_id, cartera_id)

# ── Página sync ───────────────────────────────────────────────────────────────

def pagina_sync(supabase, user_id):
    st.title("💰 Money Magnet")
    st.caption("Gestión de finanzas personales")
    st.divider()

    # ── Sección 1: Sincronizar gastos ─────────────────────────────────────────
    st.subheader("📤 Sincronizar Gastos")
    st.write("Subí el archivo exportado desde Money Manager para actualizar tus datos.")

    archivo = st.file_uploader("Seleccioná tu archivo xlsx", type=["xlsx"],
                                help="Exportá desde Money Manager: Ajustes → Respaldo → Exportar")

    if archivo:
        try:
            # Validar columnas duplicadas
            cols_originales = pd.read_excel(archivo, nrows=0).columns.tolist()
            archivo.seek(0)
            duplicadas = [c for c in set(cols_originales) if cols_originales.count(c) > 1]
            if duplicadas:
                st.warning(
                    f"⚠️ El archivo tiene columnas duplicadas: {', '.join(duplicadas)}. "
                    "Pandas las renombrará automáticamente, pero revisa que no afecte a los datos."
                )

            # Detectar cuentas disponibles
            df_preview_cuentas = pd.read_excel(archivo)
            archivo.seek(0)
            cuentas_disponibles = sorted(df_preview_cuentas["Cuentas"].dropna().unique().tolist())

            cuenta_sel = st.selectbox(
                "¿Qué cuenta corresponde a tus gastos personales?",
                cuentas_disponibles,
                index=cuentas_disponibles.index("Euros") if "Euros" in cuentas_disponibles else 0
            )

            df = procesar_xlsx(archivo, nombre_cuenta=cuenta_sel)
            st.success(f"✅ Archivo cargado: **{len(df)} registros** de cuenta '{cuenta_sel}' detectados")
            st.dataframe(df[["fecha_gasto", "categoria_consumo", "monto", "tipo"]].head(10),
                         use_container_width=True)
            st.caption(f"Mostrando 10 de {len(df)} registros")
            st.divider()
            if st.button("🔄 Sincronizar con Supabase", type="primary"):
                with st.spinner("Sincronizando..."):
                    total = sincronizar(df, supabase, user_id)
                    balance = df.apply(
                        lambda r: r["monto"] if r["tipo"] == "Ingreso" else -r["monto"], axis=1
                    ).sum()
                get_todos_gastos.clear()
                get_balance_app.clear()
                st.session_state.mostrar_saldos_post_sync = True
                st.session_state.pop("saldos_temp", None)
                st.success(f"""
                ✅ **Sincronización completada**
                - 📊 **{total:,} registros** subidos a Supabase
                - 💰 **Balance actual: €{balance:,.2f}**
                - 🕐 **{datetime.now().strftime('%d/%m/%Y %H:%M')}**
                """)
        except ValueError as e:
            st.error(f"❌ Error en el archivo: {e}")
        except Exception as e:
            st.error(f"❌ Error al sincronizar: {e}")

    if st.session_state.get("mostrar_saldos_post_sync", False):
        widget_saldos_inline(supabase, user_id)

    # ── Sección 2: Cargar presupuestos ────────────────────────────────────────
    st.divider()
    st.subheader("🎯 Cargar Presupuestos")
    st.write("Subí un CSV con el presupuesto mensual para cargarlo en Supabase.")
    st.caption("Formato requerido: columnas `categoria_consumo`, `fecha` (YYYY-MM-01), `monto`")

    csv_file = st.file_uploader("Seleccioná tu archivo CSV", type=["csv"],
                                 help="Una fila por categoría. Ingresos positivos, gastos negativos.",
                                 key="csv_presupuesto")

    if csv_file:
        try:
            df_csv = pd.read_csv(csv_file)
            columnas_req = ["categoria_consumo", "fecha", "monto"]
            faltantes = [c for c in columnas_req if c not in df_csv.columns]
            if faltantes:
                st.error(f"❌ Columnas faltantes: {faltantes}")
                return

            df_csv["fecha"] = pd.to_datetime(df_csv["fecha"]).dt.strftime("%Y-%m-%d")
            df_csv["monto"] = pd.to_numeric(df_csv["monto"], errors="coerce")
            df_csv = df_csv.dropna(subset=["monto"])
            df_csv["categoria_consumo"] = df_csv["categoria_consumo"].str.strip()

            mes_detectado = df_csv["fecha"].iloc[0][:7]
            st.success(f"✅ CSV cargado: **{len(df_csv)} categorías** para **{mes_detectado}**")
            df_preview = df_csv.copy()
            df_preview["monto"] = df_preview["monto"].apply(lambda x: f"€{x:,.2f}")
            st.dataframe(df_preview, use_container_width=True, hide_index=True)

            st.divider()
            if st.button("💾 Cargar presupuesto en Supabase", type="primary", key="btn_presupuesto"):
                with st.spinner("Cargando presupuesto..."):
                    registros = df_csv.to_dict(orient="records")
                    ok = errores = 0
                    for r in registros:
                        try:
                            supabase.table("presupuestos").upsert({
                                "categoria_consumo": r["categoria_consumo"],
                                "fecha": r["fecha"],
                                "monto": float(r["monto"]),
                                "user_id": user_id
                            }, on_conflict="categoria_consumo,fecha,user_id").execute()
                            ok += 1
                        except Exception:
                            errores += 1
                st.success(f"✅ **{ok} categorías** cargadas correctamente" +
                           (f" — ⚠️ {errores} errores" if errores > 0 else ""))
        except Exception as e:
            st.error(f"❌ Error al procesar el CSV: {e}")

    # ── Sección 3: Carga inicial xlsx cartera ─────────────────────────────────
    st.divider()
    st.subheader("💼 Carga inicial de cartera (xlsx)")
    st.caption("Usa esto solo para cargar una cartera por primera vez desde Google Sheets.")

    carteras = get_carteras(supabase, user_id)
    if not carteras:
        st.info("Primero crea una cartera en la página 💼 Cartera.")
    else:
        opciones_carteras = {c["nombre"]: c["id"] for c in carteras}
        cartera_sel = st.selectbox("Selecciona la cartera a cargar",
                                    list(opciones_carteras.keys()),
                                    key="sel_cartera_sync")
        cartera_id_sel = opciones_carteras[cartera_sel]

        archivo_cartera = st.file_uploader("Selecciona tu archivo xlsx de cartera",
                                            type=["xlsx"], key="xlsx_cartera")

        if archivo_cartera:
            try:
                df_trans, df_tickers = procesar_xlsx_cartera(archivo_cartera)
                st.success(f"✅ **{len(df_trans)} transacciones** · **{len(df_tickers)} tickers** únicos")
                st.dataframe(
                    df_trans[["ticker", "tipo", "fecha_operacion", "cantidad", "precio_entrada"]].head(10),
                    use_container_width=True
                )
                st.caption(f"Mostrando 10 de {len(df_trans)} transacciones")
                st.warning(f"⚠️ Esto **reemplazará** todas las posiciones de la cartera **{cartera_sel}**.")
                if st.button("🔄 Cargar en Supabase", type="primary", key="btn_sync_cartera"):
                    with st.spinner("Cargando cartera..."):
                        total = sincronizar_cartera(df_trans, df_tickers, supabase,
                                                     user_id, cartera_id_sel)
                        get_cartera.clear()
                        get_tickers_sin_sector.clear()
                    st.success(f"✅ **{total} transacciones** cargadas en '{cartera_sel}'")
            except Exception as e:
                st.error(f"❌ Error: {e}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not login_page():
        return

    user_id = get_user_id()
    supabase = get_supabase_auth()

    st.sidebar.title("📱 Navegación")

    if st.sidebar.button("🚪 Cerrar sesión"):
        st.session_state.user = None
        st.session_state.access_token = None
        st.rerun()

    pagina = st.sidebar.radio(
        "Ir a:",
        ["📊 Dashboard", "📈 Histórico", "🔍 Detalle", 
         "💳 Bancos", "🔮 Proyección", "💼 Cartera", "🤖 Prompt IA", "📤 Sincronizar"],
        index=0
    )

    st.sidebar.divider()
    balance_app = get_balance_app(supabase, user_id)
    df_saldos, ultima_fecha = get_saldos_actuales(supabase, user_id)

    if df_saldos.empty:
        st.sidebar.info("💳 Sin saldos registrados")
    else:
        total_bancos = df_saldos["monto"].sum()
        diferencia = abs(balance_app - total_bancos)
        if diferencia <= 0.01:
            st.sidebar.success(f"✅ Data cuadrada\n\n€{total_bancos:,.2f}")
        else:
            st.sidebar.warning(f"⚠️ Revisar saldos\n\nDif: €{diferencia:,.2f}")

    if pagina == "📊 Dashboard":
        pagina_dashboard(supabase, user_id)
    elif pagina == "📈 Histórico":
        pagina_historico(supabase, user_id)
    elif pagina == "🔍 Detalle":
        pagina_detalle(supabase, user_id)
    elif pagina == "💳 Bancos":
        pagina_bancos(supabase, user_id)
    elif pagina == "🔮 Proyección":
        pagina_proyeccion(supabase, user_id)
    elif pagina == "📤 Sincronizar":
        pagina_sync(supabase, user_id)
    elif pagina == "💼 Cartera":
        pagina_cartera(supabase, user_id)
    elif pagina == "🤖 Prompt IA":
        pagina_prompt(supabase, user_id)

if __name__ == "__main__":
    main()